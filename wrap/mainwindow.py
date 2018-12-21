from gl import *
from ui import *
from wrap.business import Business
from openpyxl import Workbook
from PyQt5.QtWidgets import QMainWindow,QDialog,QTableWidgetItem,QAbstractItemView,QListWidgetItem,QMenu,QAction,QMessageBox,QFileDialog,QMessageBox,QInputDialog,QTreeWidgetItem,QFileDialog
from PyQt5.QtCore import QDate,Qt
from PyQt5.QtGui import QIcon,QCursor,QBrush

class FilterDialog(QDialog, Ui_FilterDialog):
    def __init__(self, parent=None):
        QDialog.__init__(self, parent)
        self.setupUi(self)

        self.twFilter.setContextMenuPolicy(Qt.CustomContextMenu)
        self.twFilter.customContextMenuRequested.connect(self.twFilterMenu)
        self.twFilter.horizontalHeader().setStyleSheet("QHeaderView::section{background:skyblue;}")
        self.twFilter.verticalHeader().setStyleSheet("QHeaderView::section{background:skyblue;}")
        #self.twFilter.setEditTriggers(QAbstractItemView.NoEditTriggers);
        self.twFilter.itemDoubleClicked.connect(self.tableFilterItemEdit)
        self.menuTwFilter = QMenu(self)

        self.actDel = QAction(self)
        self.actDel.setText('删除')
        self.menuTwFilter.addAction(self.actDel)
        self.actDel.triggered.connect(self.actDelClicked)
        self.actDelCol = QAction(self)
        self.actDelCol.setText('删除列')
        self.menuTwFilter.addAction(self.actDelCol)
        self.actDelCol.triggered.connect(self.actDelColClicked)
        self.actClear = QAction(self)
        self.actClear.setText('清空')
        self.menuTwFilter.addAction(self.actClear)
        self.actClear.triggered.connect(self.actClearClicked)
        self.actFlush = QAction(self)
        self.actFlush.setText('刷新')
        self.menuTwFilter.addAction(self.actFlush)
        self.actFlush.triggered.connect(self.actFlushClicked)

        self.btnAddField.clicked.connect(self.btnAddFieldClicked)

        self.zhHead = []
        self.enHead = []
        self.data = {}
        self.fullHead = None

    def twFilterMenu(self):
        self.menuTwFilter.popup(QCursor.pos())

    def clear(self):
        self.twFilter.clear()
        self.zhHead.clear()
        self.enHead.clear()
        self.data.clear()

    def add(self, zhHead, value):
        index = self.fullHead[1].index(zhHead)
        enHead = self.fullHead[0][index]
        typ = self.fullHead[2][index]

        if zhHead not in self.zhHead:
            self.zhHead.append(zhHead)
            self.enHead.append(enHead)
            self.data[zhHead] = []

        try:
            if value!=None and value not in self.data[zhHead]:
                if value=='' or value=='空':
                    value = '空'
                elif typ == 'int':
                    value = int(value)
                elif typ == 'double':
                    value = float(value)
                self.data[zhHead].append(value)
        except:
            err = '内容类型不匹配'
            GL.LOG.error(err)
            QMessageBox.critical(self, 'Error', err)

        self.flushTable()

    def actClearClicked(self):
        self.clear()

    def actFlushClicked(self):
        self.flushTable()

    def actDelClicked(self):
        it = self.twFilter.currentItem()
        if it != None:
            r = it.row()
            c = it.column()
            del self.data[self.zhHead[c]][r]
            self.twFilter.setItem(it.row(), it.column(), None)
            self.flushTable()

    def actDelColClicked(self):
        c = self.twFilter.currentColumn()
        it = self.twFilter.horizontalHeaderItem(c)
        if it != None:
            del self.data[self.zhHead[c]]
            del self.enHead[c]
            del self.zhHead[c]
            self.flushTable()

    def sql(self):
        if len(self.zhHead) == 0:
            return None
        sql = ''
        for n in range(0, len(self.zhHead)):
            if len(self.data[self.zhHead[n]]) == 0:
                continue
            if '空' in self.data[self.zhHead[n]]:
                sql += '(%s is null or %s in %s) and ' % (self.enHead[n],self.enHead[n],self.data[self.zhHead[n]])
            else:
                sql += '%s in %s and ' % (self.enHead[n],self.data[self.zhHead[n]])
        sql = sql.rstrip(' and ')
        sql = sql.replace('[', '(')
        sql = sql.replace(']', ')')
        sql = sql.replace(', \'空\'', '')
        return sql

    def flushTable(self):
        self.twFilter.clear()
        self.twFilter.setColumnCount(20)
        self.twFilter.setRowCount(15)
        for r in range(0,self.twFilter.rowCount()):
            for c in range(0,self.twFilter.columnCount()):
                it = QTableWidgetItem('')
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                self.twFilter.setItem(r, c, it)
        self.twFilter.setHorizontalHeaderLabels(self.zhHead)
        for c in range(0, len(self.zhHead)):
            for r in range(0, len(self.data[self.zhHead[c]])):
                value = self.data[self.zhHead[c]][r]
                it = QTableWidgetItem(str(value))
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                self.twFilter.setItem(r, c, it)

    def flushCombobox(self, head):
        self.fullHead = head
        self.cmbField.clear()
        self.cmbField.addItems(head[1])

    def tableFilterItemEdit(self, item):
        if item == None:
            return
        r = item.row()
        c = item.column()
        if c >= len(self.zhHead):
            return
        old = item.text()
        (new, ok) = QInputDialog.getText(self,'编辑表格','输入新内容：', text=old)
        if new == '':
            new = '空'
        if ok and new!=old:
            if r < len(self.data[self.zhHead[c]]):
                del self.data[self.zhHead[c]][r]
            self.add(self.zhHead[c], new)
            self.flushTable()

    def btnAddFieldClicked(self):
        field = self.cmbField.currentText()
        if field not in self.zhHead:
            self.add(field, None)

class MainWindow(QMainWindow, Ui_MainWindow):

    def __init__(self, parent=None):
        QMainWindow.__init__(self, parent)
        self.setupUi(self)
        GL.LOG = getLogger('WanboLoger', 'logs', 'console.log')
        GL.LOG.info('主程序启动')
        self.bus = Business()
        self.init()

    def closeEvent(self, event):
        if self.bus != None:
            del self.bus
        GL.LOG.info('主程序关闭')

    def actQryTestClicked(self):
        pass

    def init(self):
        #数据管理(Query)页面的表格
        self.twQuery.setContextMenuPolicy(Qt.CustomContextMenu)
        self.twQuery.customContextMenuRequested.connect(self.tableQueryMenu)
        self.twQuery.horizontalHeader().setStyleSheet("QHeaderView::section{background:skyblue;}")
        self.twQuery.verticalHeader().setStyleSheet("QHeaderView::section{background:skyblue;}")
        self.twQuery.itemDoubleClicked.connect(self.tableQueryItemEdit)

        self.trQuery.setColumnCount(1)
        self.trQuery.setHeaderHidden(True)
        t = QTreeWidgetItem(self.trQuery, ['表格管理',])
        for k,v in self.bus.tables().items():
            QTreeWidgetItem(t, [k,])
        self.trQuery.expandAll()
        self.trQuery.currentItemChanged.connect(self.treeQueryItemActivated)
        self.trQuery.sortItems(0, Qt.AscendingOrder)

        #统计汇总(stats)页面的表格
        self.twStats.setContextMenuPolicy(Qt.CustomContextMenu)
        self.twStats.customContextMenuRequested.connect(self.tableStatsMenu)
        self.twStats.horizontalHeader().setStyleSheet("QHeaderView::section{background:skyblue;}")
        self.twStats.verticalHeader().setStyleSheet("QHeaderView::section{background:skyblue;}")
        #self.twStats.itemDoubleClicked.connect(self.tableStatsItemEdit)

        self.trStats.setColumnCount(1)
        self.trStats.setHeaderHidden(True)
        t = QTreeWidgetItem(self.trStats, ['统计汇总',])
        for k in self.bus.stats():
            QTreeWidgetItem(t, [k,])
        self.trStats.expandAll()
        self.trStats.currentItemChanged.connect(self.treeStatsItemActivated)
        self.trStats.sortItems(0, Qt.AscendingOrder)

        #操作(Job)页面的表格
        self.twJob.setContextMenuPolicy(Qt.CustomContextMenu)
        self.twJob.customContextMenuRequested.connect(self.tableJobMenu)
        self.twJob.horizontalHeader().setStyleSheet("QHeaderView::section{background:skyblue;}")
        self.twJob.verticalHeader().setStyleSheet("QHeaderView::section{background:skyblue;}")
        self.twJob.itemDoubleClicked.connect(self.tableJobItemEdit)

        #数据管理(Query)页面的右键菜单
        self.menuTableQuery = QMenu(self)

        self.actQryFilter = QAction(self)
        self.actQryFilter.setText('筛选')
        self.menuTableQuery.addAction(self.actQryFilter)
        self.actQryFilter.triggered.connect(self.actQryFilterClicked)

        self.actQryRelateAccount = QAction(self)
        self.actQryRelateAccount.setText('关联账款')
        self.menuTableQuery.addAction(self.actQryRelateAccount)
        self.actQryRelateAccount.triggered.connect(self.actQryRelateAccountClicked)

        self.actQryRelateDetail = QAction(self)
        self.actQryRelateDetail.setText('关联明细')
        self.menuTableQuery.addAction(self.actQryRelateDetail)
        self.actQryRelateDetail.triggered.connect(self.actQryRelateDetailClicked)

        self.actQryTest = QAction(self)
        self.actQryTest.setText('测试')
        self.menuTableQuery.addAction(self.actQryTest)
        self.actQryTest.triggered.connect(self.actQryTestClicked)

        self.actQryAdvFilter = QAction(self)
        self.actQryAdvFilter.setText('添加到高级筛选')
        self.menuTableQuery.addAction(self.actQryAdvFilter)
        self.actQryAdvFilter.triggered.connect(self.actQryAdvFilterClicked)

        #操作(Job)页面的右键菜单
        self.menuTableJob = QMenu(self)

        self.actJobRelate = QAction(self)
        self.actJobRelate.setText('确认关联')
        self.menuTableJob.addAction(self.actJobRelate)
        self.actJobRelate.triggered.connect(self.actJobRelateClicked)

        #统计汇总页面的右键菜单
        self.menuTableStats = QMenu(self)

        self.actStatsExport = QAction(self)
        self.actStatsExport.setText('导出')
        self.menuTableStats.addAction(self.actStatsExport)
        self.actStatsExport.triggered.connect(self.actStatsExportClicked)

        #其他控件关联
        self.btnRefresh.clicked.connect(self.btnRefreshClicked)
        self.btnJobRefresh.clicked.connect(self.btnJobRefreshClicked)
        self.btnAdd.clicked.connect(self.btnAddClicked)
        self.btnJobClose.clicked.connect(self.btnJobCloseClicked)
        self.edtFilter.textChanged.connect(self.edtFilterChanged)
        self.edtFilterJob.textChanged.connect(self.edtFilterJobChanged)
        self.btnBrowse.clicked.connect(self.btnBrowseClicked)
        self.btnImport.clicked.connect(self.btnImportClicked)
        self.btnClearMsg.clicked.connect(self.btnClearMsgClicked)
        self.btnAdvFilter.clicked.connect(self.btnAdvFilterClicked)

        #更多需要初始化的内容
        self.dlg = FilterDialog(self)
        self.edtFile.setFocusPolicy(Qt.NoFocus)
        self.txtLoadMsg.setFocusPolicy(Qt.NoFocus)
        self.jobTabIndex = 3
        self.tab.removeTab(self.jobTabIndex)   #默认隐藏job标签页
        self.tableQuery = None
        self.tableQueryZh = None
        self.tableQueryData = None
        self.tableQueryHead = None
        self.tableJob = None
        self.tableJobZh = None
        self.tableJobData = None
        self.tableJobHead = None
        self.tableDst = None
        self.twDst = None
        self.twDstHead = None
        self.itemDst = None
        self.txtDst = None
        self.isAdding = False

    def btnAdvFilterClicked(self):
        self.dlg.show()

    def actQryAdvFilterClicked(self):
        it = self.twQuery.currentItem()
        if it!=None and it.text()!=None:
            itHead = self.twQuery.horizontalHeaderItem(it.column())
            zhHead = itHead.text()
            self.dlg.add(zhHead, it.text())

    def btnClearMsgClicked(self):
        self.txtLoadMsg.clear()

    def btnBrowseClicked(self):
        fn,ft = QFileDialog.getOpenFileName(self, '选择导入文件', 'D:\MYC\data\wanbo', 'Excel Files (*.xlsx)')
        self.edtFile.setText(fn)

    def btnImportClicked(self):
        wb = self.bus.loadExcel(self.edtFile.text())
        for ws in wb:
            if ws.title=='现金账户1明细账' and self.cb1.isChecked():
                source = '平安银行（姚洋）'
                (r1,r2) = self.bus.ReadBalanceData(ws, source)
                self.txtLoadMsg.append('平安银行（姚洋），导入：成功 %d, 失败 %d' % (r1,r2))
            elif ws.title=='现金账户2明细账' and self.cb2.isChecked():
                source = '平安银行（李昱平）'
                (r1,r2) = self.bus.ReadBalanceData(ws, source)
                self.txtLoadMsg.append('平安银行（李昱平），导入：成功 %d, 失败 %d' % (r1,r2))
            elif ws.title=='银行明细账' and self.cb3.isChecked():
                source = '建设银行（基本户）'
                (r1,r2) = self.bus.ReadBalanceData(ws, source)
                self.txtLoadMsg.append('建设银行（基本户），导入：成功 %d, 失败 %d' % (r1,r2))
            elif ws.title=='应收账款汇总表' and self.cb4.isChecked():
                (r1,r2) = self.bus.ReadAccountData(ws)
                self.txtLoadMsg.append('应收账款，导入：成功 %d, 失败 %d' % (r1,r2))
            elif ws.title=='合同明细' and self.cb5.isChecked():
                (r1,r2) = self.bus.ReadContractData(ws)
                self.txtLoadMsg.append('合同明细，导入：成功 %d, 失败 %d' % (r1,r2))
            elif ws.title=='开票明细表' and self.cb6.isChecked():
                (r1,r2) = self.bus.ReadInvoiceData(ws)
                self.txtLoadMsg.append('开票明细，导入：成功 %d, 失败 %d' % (r1,r2))
            else:
                pass

    def treeQueryItemActivated(self, itemNew, itemOld):
        self.cbFilter.setChecked(False)
        self.dlg.clear()
        self.fillTableQuery(itemNew.text(0))

    def tableQueryItemEdit(self, item):
        if self.isAdding:
            return
        r = item.row()
        c = item.column()
        if c == 0:
            self.statusbar.showMessage('编号不可修改！', 5000)
            return
        old = item.text()
        (new, ok) = QInputDialog.getText(self,'编辑表格','输入新内容：', text=old)
        if ok and new!=old:
            id_it = self.twQuery.item(r, 0)
            id_enHead = self.tableQueryHead[0][0]
            id_value = int(id_it.text())
            enHead = self.tableQueryHead[0][c]
            zhHead = self.tableQueryHead[1][c]
            tp = self.tableQueryHead[2][c]
            if self.bus.updateTableById(self.tableQuery, enHead, tp, new, id_enHead, id_value):
                item.setText(new)
                GL.LOG.info('编辑表格(%s), id(%d)的(%s)由(%s)改为(%s).' % (self.tableQuery,id_value,zhHead,old,new))
            else:
                QMessageBox.critical(self, 'Error', '失败！')

    def tableJobItemEdit(self, item):
        r = item.row()
        c = item.column()
        if c == 0:
            self.statusbar.showMessage('编号不可修改！', 5000)
            return
        old = item.text()
        (new, ok) = QInputDialog.getText(self,'编辑表格','输入新内容：', text=old)
        if ok and new!=old:
            id_it = self.twJob.item(r, 0)
            id_enHead = self.tableJobHead[0][0]
            id_value = int(id_it.text())
            enHead = self.tableJobHead[0][c]
            zhHead = self.tableJobHead[1][c]
            tp = self.tableJobHead[2][c]
            if self.bus.updateTableById(self.tableJob, enHead, tp, new, id_enHead, id_value):
                item.setText(new)
                GL.LOG.info('编辑表格(%s), id(%d)的(%s)由(%s)改为(%s).' % (self.tableJob,id_value,zhHead,old,new))
            else:
                QMessageBox.critical(self, 'Error', '失败！')

    def btnJobCloseClicked(self):
        self.resetTabJob()

    def tableQueryMenu(self, pos):
        self.actQryRelateAccount.setVisible(False)
        self.actQryRelateDetail.setVisible(False)
        if self.twQuery.currentItem() != None:
            self.actQryFilter.setVisible(True)
            if self.tableQueryZh == '应收账款':
                self.actQryRelateDetail.setVisible(True)
            elif self.tableQueryZh == '收支明细':
                self.actQryRelateAccount.setVisible(True)
        self.menuTableQuery.popup(QCursor.pos())

    def tableJobMenu(self, pos):
        self.actJobRelate.setVisible(False)
        if self.tableJobZh=='应收账款' or self.tableJobZh=='收支明细':
            self.actJobRelate.setVisible(True)
        self.menuTableJob.popup(QCursor.pos())

    def tableStatsMenu(self, pos):
        self.menuTableStats.popup(QCursor.pos())

    def actQryFilterClicked(self):
        item = self.twQuery.currentItem()
        if item == None:
            return
        self.edtFilter.setText(item.text())

    def resetTabJob(self):
        self.twJob.clear()
        self.edtFilterJob.clear()
        self.tab.removeTab(self.jobTabIndex)
        self.tab.setCurrentIndex(0)

    def actQryRelateAccountClicked(self):
        item = self.twQuery.currentItem()
        if item == None:
            return
        self.resetTabJob()
        table_zh = '应收账款'
        self.tab.insertTab(self.jobTabIndex, self.tabJob, '操作-%s'%table_zh)
        self.tab.setCurrentIndex(self.jobTabIndex)
        self.fillTableJob(table_zh)
        self.edtFilterJob.setText(item.text())

        self.tableDst = self.tableQuery
        self.twDst = self.twQuery
        self.tableDstHead = self.tableQueryHead
        self.itemDst = self.twQuery.item(item.row(), 1)
        self.txtDst = None

    def actQryRelateDetailClicked(self):
        item = self.twQuery.currentItem()
        if item == None:
            return
        self.resetTabJob()
        table_zh = '收支明细'
        self.tab.insertTab(self.jobTabIndex, self.tabJob, '操作-%s'%table_zh)
        self.tab.setCurrentIndex(self.jobTabIndex)
        self.fillTableJob(table_zh)
        self.edtFilterJob.setText(item.text())
        #操作页加载后才能取到tableJob的各项数据
        self.tableDst = self.tableJob
        self.twDst = self.twJob
        self.tableDstHead = self.tableJobHead
        self.itemDst = None
        self.txtDst = self.twQuery.item(item.row(),0).text()

    def actJobRelateClicked(self):
        item = self.twJob.currentItem()
        if item == None:
            return
        if self.itemDst!=None and self.txtDst==None:
            self.txtDst = self.twJob.item(item.row(),0).text()
        elif self.itemDst==None and self.txtDst!=None:
            self.itemDst = self.twJob.item(item.row(),1)
        else:
            QMessageBox.critical(self, 'Error', '关联时数据有误, 请重试！')
            return
        r = item.row()
        c = 1   #关联就是修改明细表的第二列
        id_it = self.twDst.item(r, 0)
        id_enHead = self.tableDstHead[0][0]
        id_value = int(id_it.text())
        enHead = self.tableDstHead[0][c]
        zhHead = self.tableDstHead[1][c]
        tp = self.tableDstHead[2][c]
        old = self.itemDst.text()
        if self.bus.updateTableById(self.tableDst, enHead, tp, self.txtDst, id_enHead, id_value):
            GL.LOG.info('编辑表格(%s), id(%d)的(%s)由(%s)改为(%s).' % (self.tableDst,id_value,zhHead,old,self.txtDst))
            self.itemDst.setText(self.txtDst)
            self.twDst.setCurrentItem(self.itemDst)
        else:
            QMessageBox.critical(self, 'Error', '失败！')

        self.tableDst = None
        self.twDst = None
        self.twDstHead = None
        self.itemDst = None
        self.txtDst = None

    def fillTableQuery(self, table):
        if table not in self.bus.tables():
            return
        self.isAdding = False
        condition = None
        if self.cbFilter.isChecked():
            condition = self.dlg.sql()
        self.tableQuery = self.bus.tables()[table]
        self.tableQueryZh = table
        self.tableQueryData = self.bus.selectTableData(self.tableQuery, condition)
        self.tableQueryHead = self.bus.selectTableHead(self.tableQuery)
        self.fillTable(self.twQuery, self.tableQuery, self.tableQueryData, self.tableQueryHead)
        self.dlg.flushCombobox(self.tableQueryHead)

    def fillTableJob(self, table):
        self.tableJob = self.bus.tables()[table]
        self.tableJobZh = table
        self.tableJobData = self.bus.selectTableData(self.tableJob)
        self.tableJobHead = self.bus.selectTableHead(self.tableJob)
        self.fillTable(self.twJob, self.tableJob, self.tableJobData, self.tableJobHead)

    def btnRefreshClicked(self):
        self.fillTableQuery(self.tableQueryZh)

    def btnJobRefreshClicked(self):
        self.fillTableJob(self.tableJobZh)

    def edtFilterChanged(self, txt):
        if txt == '':
            for r in range(0,self.twQuery.rowCount()):
                self.twQuery.setRowHidden(r, False)
            return
        if len(txt) < 2:
            return
        current = self.twQuery.currentItem()
        items = self.twQuery.findItems(txt, Qt.MatchContains)
        showRows = []
        for it in items:
            if it.row() not in showRows:
                showRows.append(it.row())
        self.twQuery.setVisible(False)
        for i in range(0,self.twQuery.rowCount()):
            if i not in showRows:
                self.twQuery.setRowHidden(i, True)
            else:
                self.twQuery.setRowHidden(i, False)
        self.twQuery.setCurrentItem(current)
        self.twQuery.setVisible(True)

    def edtFilterJobChanged(self, txt):
        if txt == '':
            for r in range(0,self.twJob.rowCount()):
                self.twJob.setRowHidden(r, False)
            return
        if len(txt) < 2:
            return
        current = self.twJob.currentItem()
        items = self.twJob.findItems(txt, Qt.MatchContains)
        showRows = []
        for it in items:
            if it.row() not in showRows:
                showRows.append(it.row())
        self.twJob.setVisible(False)
        for i in range(0,self.twJob.rowCount()):
            if i not in showRows:
                self.twJob.setRowHidden(i, True)
            else:
                self.twJob.setRowHidden(i, False)
        self.twJob.setCurrentItem(current)
        self.twJob.setVisible(True)

    def fillTable(self, tw, table, data, head):
        tw.clear()
        tw.setColumnCount(len(head[1]))
        tw.setRowCount(len(data))
        tw.setHorizontalHeaderLabels(head[1])
        self.btnAdd.setText('新增')

        tw.setVisible(False)
        for r in range(0,tw.rowCount()):
            for c in range(0,tw.columnCount()):
                tmp = data[r][head[0][c]]
                if tmp == None:
                    tmp = ''
                it = QTableWidgetItem(str(tmp))
                tw.setItem(r,c,it)
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                tw.setRowHidden(r, False)
        tw.setVisible(True)

    def btnAddClicked(self):
        if self.tableQueryZh not in self.bus.tables():
            return
        if self.btnAdd.text() == '新增':
            self.isAdding = True
            self.twQuery.clearContents()
            self.twQuery.setRowCount(1)
            self.btnAdd.setText('确认新增')
        else:
            datas = []
            r = 0
            for c in range(0,self.twQuery.columnCount()):
                it = self.twQuery.item(r, c)
                txt = None
                #表格为空或都是空格就以None处理
                if it != None:
                    txt = it.text()
                    if txt.strip() == '':
                        txt = 'NULL'
                else:
                    txt = 'NULL'
                datas.append(txt)
            if self.bus.insertTable(self.tableQueryZh, datas):
                self.btnAdd.setText('新增')
                self.btnRefreshClicked()
            else:
                QMessageBox.critical(self, 'Error', '失败！')

    def treeStatsItemActivated(self, itemNew, itemOld):
        if itemNew.text(0) == '应收账款统计':
            tw = self.twStats
            tw.clear()
            tw.setColumnCount(7)
            tw.setRowCount(13)
            ret = self.bus.getAccountStats()
            if ret['今年合同额'] != 0.0:
                ret['今年合同欠款率'] = ret['今年未收款额'] / ret['今年合同额']
            else:
                ret['今年合同欠款率'] = 0.0
            if ret['历年合同额'] != 0.0:
                ret['历年合同欠款率'] = ret['历年未收款额'] / ret['历年合同额']
            else:
                ret['历年合同欠款率'] = 0.0
            if ret['合同总额'] != 0.0:
                ret['总欠款率'] = ret['未收款总额'] / ret['合同总额']
            else:
                ret['总欠款率'] = 0.0
            struct = self.bus.statsAccount()
            for k,v in ret.items():
                if k not in struct:
                    continue
                if v == None:
                    v = 0.0
                it1 = QTableWidgetItem(str(k))
                it1.setFlags(it1.flags() & ~Qt.ItemIsEditable)
                it1.setBackground(QBrush(Qt.lightGray))
                if struct[k]['form'] == '百分比':
                    it2 = QTableWidgetItem('%.2f%%' % (v*100))
                else:
                    it2 = QTableWidgetItem(str(v))
                it2.setFlags(it2.flags() & ~Qt.ItemIsEditable)
                row = struct[k]['row']
                col = struct[k]['column']
                tw.setItem(row, col, it1)
                tw.setItem(row+1, col, it2)
        elif itemNew.text(0) == '开票统计':
            tw = self.twStats
            tw.clear()
            heads = self.bus.statsInvoice()
            ret = self.bus.getInvoiceStats()
            tw.setColumnCount(len(heads))
            tw.setRowCount(len(ret))
            tw.setHorizontalHeaderLabels(heads)
            row = 0
            for tmp in ret:
                for k,v in tmp.items():
                    if v == None:
                        v = 0.0
                    col = heads.index(k)
                    it = QTableWidgetItem(str(v))
                    it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                    tw.setItem(row, col, it)
                it = QTableWidgetItem(str(v))
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                tw.setItem(row, col, it)
                row += 1
        elif itemNew.text(0) == '收支统计':
            tw = self.twStats
            tw.clear()
            heads = self.bus.statsBalance()
            (ret, rowCount) = self.bus.getBalanceStats()
            tw.setColumnCount(len(heads))
            tw.setRowCount(rowCount)
            tw.setHorizontalHeaderLabels(heads)
            row = 0
            for l in ret:
                for k1,v1 in l.items():
                    it = QTableWidgetItem(str(k1))
                    it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                    col = 0
                    tw.setItem(row, col, it)
                    for k2,v2 in v1.items():
                        it = QTableWidgetItem(str(k2))
                        it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                        col = 1
                        tw.setItem(row, col, it)
                        for k3,v3 in v2.items():
                            it = QTableWidgetItem(str(v3))
                            it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                            col = heads.index(k3)
                            tw.setItem(row, col, it)
                        row += 1
                    row += 1
        elif itemNew.text(0) == '费用统计':
            tw = self.twStats
            tw.clear()
            heads = self.bus.statsCost()
            (ret,count) = self.bus.getCostStats()
            tw.setColumnCount(len(heads))
            tw.setRowCount(count)
            tw.setHorizontalHeaderLabels(heads)
            #总费用
            row = 0
            col = 0
            it = QTableWidgetItem('总费用汇总')
            it.setFlags(it.flags() & ~Qt.ItemIsEditable)
            it.setBackground(QBrush(Qt.lightGray))
            tw.setItem(row, col, it)
            for p in ret['总费用汇总']['费用']:
                mon = p['月份']
                cost = p['费用']
                if cost == None:
                    cost = 0.0
                it = QTableWidgetItem(str(cost))
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                it.setBackground(QBrush(Qt.lightGray))
                col = heads.index(mon)
                tw.setItem(row, col, it)
            #一级和二级类目
            row = 1
            for class1,v1 in ret.items():
                if class1 == '总费用汇总':
                    continue
                #一级类目
                it = QTableWidgetItem(str(class1))
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                it.setBackground(QBrush(Qt.lightGray))
                col = 0
                tw.setItem(row, col, it)
                #一级类目的费用
                if '费用' in v1:
                    for p1 in v1['费用']:
                        mon = p1['月份']
                        cost = p1['费用']
                        if cost == None:
                            cost = 0.0
                        it = QTableWidgetItem(str(cost))
                        it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                        it.setBackground(QBrush(Qt.lightGray))
                        col = heads.index(mon)
                        tw.setItem(row, col, it)
                #二级类目
                if '二级类目' in v1:
                    for class2,v2 in v1['二级类目'].items():
                        it = QTableWidgetItem(str(class2))
                        it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                        row += 1
                        col = 1
                        tw.setItem(row, col, it)
                        #二级类目的费用
                        for p2 in v2:
                            mon = p2['月份']
                            cost = p2['费用']
                            if cost == None:
                                cost = 0.0
                            it = QTableWidgetItem(str(cost))
                            it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                            col = heads.index(mon)
                            tw.setItem(row, col, it)
                row += 1
            #处理空表格
            for r in range(0, tw.rowCount()):
                for c in range(0, tw.columnCount()):
                    it = tw.item(r, c)
                    it_0 = tw.item(r, 0)
                    if it == None:
                        if c==0 or c==1:
                            it = QTableWidgetItem('')
                        else:
                            it = QTableWidgetItem(str(0.0))
                        if it_0!=None and it_0.text()!='':
                            it.setBackground(QBrush(Qt.lightGray))
                        it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                        tw.setItem(r, c, it)
        else:
            pass


    def actStatsExportClicked(self):
        wb = Workbook()

        ws = wb.active
        ws.title = '费用统计'
        heads = self.bus.statsCost()
        (ret,count) = self.bus.getCostStats()
        ws.append(heads)
        row = 2
        col = 1
        ws.cell(row=row,column=col).value = '总费用汇总'
        for p in ret['总费用汇总']['费用']:
            mon = p['月份']
            cost = p['费用']
            if cost == None:
                cost = 0.0
            col = heads.index(mon) + 1
            ws.cell(row=row,column=col).value = cost
        #一级和二级类目
        row = 3
        for class1,v1 in ret.items():
            if class1 == '总费用汇总':
                continue
            #一级类目
            col = 1
            ws.cell(row=row,column=col).value = class1
            #一级类目的费用
            if '费用' in v1:
                for p1 in v1['费用']:
                    mon = p1['月份']
                    cost = p1['费用']
                    if cost == None:
                        cost = 0.0
                    col = heads.index(mon) + 1
                    ws.cell(row=row,column=col).value = cost
            #二级类目
            if '二级类目' in v1:
                for class2,v2 in v1['二级类目'].items():
                    row += 1
                    col = 2
                    ws.cell(row=row,column=col).value = class2
                    #二级类目的费用
                    for p2 in v2:
                        mon = p2['月份']
                        cost = p2['费用']
                        if cost == None:
                            cost = 0.0
                        col = heads.index(mon) + 1
                        ws.cell(row=row,column=col).value = cost
            row += 1

        ws = wb.create_sheet('开票统计')
        heads = self.bus.statsInvoice()
        ret = self.bus.getInvoiceStats()
        ws.append(heads)
        row = 2
        for l in ret:
            for k,v in l.items():
                col = heads.index(k) + 1
                ws.cell(row=row,column=col).value = v
            row += 1

        ws = wb.create_sheet('收支统计')
        heads = self.bus.statsBalance()
        (ret, rowCount) = self.bus.getBalanceStats()
        ws.append(heads)
        row = 2
        for l in ret:
            for k1,v1 in l.items():
                col = 1
                ws.cell(row=row,column=col).value = k1
                for k2,v2 in v1.items():
                    col = 2
                    ws.cell(row=row,column=col).value = k2
                    for k3,v3 in v2.items():
                        col = heads.index(k3) + 1
                        ws.cell(row=row,column=col).value = v3
                    row += 1
                row += 1

        ws = wb.create_sheet('应收账款统计')
        ret = self.bus.getAccountStats()
        if ret['今年合同额'] != 0.0:
            ret['今年合同欠款率'] = ret['今年未收款额'] / ret['今年合同额']
        else:
            ret['今年合同欠款率'] = 0.0
        if ret['历年合同额'] != 0.0:
            ret['历年合同欠款率'] = ret['历年未收款额'] / ret['历年合同额']
        else:
            ret['历年合同欠款率'] = 0.0
        if ret['合同总额'] != 0.0:
            ret['总欠款率'] = ret['未收款总额'] / ret['合同总额']
        else:
            ret['总欠款率'] = 0.0
        struct = self.bus.statsAccount()
        for k,v in ret.items():
            if k not in struct:
                continue
            if v == None:
                v = 0.0
            cell1 = k
            cell2 = None
            if struct[k]['form'] == '百分比':
                cell2 = '%.2f%%' % (v*100)
            else:
                cell2 = v
            row = struct[k]['row'] + 1
            col = struct[k]['column'] + 1
            ws.cell(row=row,column=col).value = cell1
            ws.cell(row=row+1,column=col).value = cell2

        fn,ft = QFileDialog.getSaveFileName(self, '保存文件', 'D:\MYC\data\wanbo', 'Excel Files (*.xlsx)')
        wb.save(fn)

